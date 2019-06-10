import openpyxl
import mysql.connector

class ConexaoMySQL():
    
    def __init__(self):
        mysql.connector.connect()
        self.__mydb = mysql.connector.connect(
            host="35.237.186.164",
            port="3306",
            user="root",
            passwd="stagiopbd2019",
            database="stagiopbd",
            auth_plugin='mysql_native_password'
        )
    
    def getData(self, sql):
        mycursor = self.__mydb.cursor()
        mycursor.execute(sql)
        myresult = mycursor.fetchall()

        return myresult

    def setData(self, sql):
        mycursor = self.__mydb.cursor()
        mycursor.execute(sql)
        self.__mydb.commit()
        print(mycursor.rowcount, "record inserted.")

    def savePatient(self, cpf, nome, dataNasc, genero, tipoSang, email):
        pat_cpf = cpf
        pat_name = nome
        pat_birthdate = dataNasc
        pat_gender = genero
        pat_bloodtype = tipoSang
        pat_email = email
        sql = "INSERT INTO patient(pat_cpf, pat_name, pat_gender, pat_bloodtype, pat_birthdate, pat_email) VALUES(\'"+pat_cpf+"', \'"+pat_name+"', \'"+pat_gender+"', \'"+pat_bloodtype+"', \'"+pat_birthdate+"', \'"+pat_email+"')"
        self.setData(sql)

class Workbook():
    
    def __init__(self):
        self.__con = ConexaoMySQL()

    def insertPatient(self):
        book = openpyxl.load_workbook("1_Paciente/etc/patient_mescled.xlsx", True)
        sheet = book["patient"]
        #lista = []
        for i in range(2, sheet.max_row + 1):
            cpf = sheet["A"+str(i)].value
            nome = sheet["B"+str(i)].value
            dataNasc = sheet["C"+str(i)].value
            genero = sheet["D"+str(i)].value
            tipoSang = sheet["E"+str(i)].value
            email = sheet["F"+str(i)].value
            #print(i)
            #lista.append([cpf, nome, dataNasc, gender, tipoSang, email])
            self.__con.savePatient(cpf, nome, dataNasc, genero, tipoSang, email)

if __name__ == '__main__':
    wbook = Workbook()
    wbook.insertPatient()
    