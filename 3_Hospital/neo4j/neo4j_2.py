from py2neo import Graph, Node, Path, Relationship
import openpyxl
# References on https://py2neo.org/2.0/essentials.html

class ConectarNeo4J():
    __graph = Graph()

    def __init__(self):
        self.__graph = Graph("bolt://35.237.186.164:7687", bolt=True, auth=("neo4j", "stagiopbd"))

    def limpar(self):
        tx = self.__graph.begin()
        tx.run("MATCH (n) DETACH DELETE n")
        tx.commit()

    def returnDataset(self, cypher):
        tx = self.__graph.begin()
        for record in tx.run(cypher):
            print(record["n"])

    def Area(self, name, city, state):
        tx = self.__graph.begin()

        area = Node("Area")
        area["name"] = name
        area["city"] = city
        area["state"] = state
        tx.create(area)
        tx.commit()

    def Zipcode(self, code, street, lat, long):
        tx = self.__graph.begin()

        zipcode = Node("Zipcode")
        zipcode["code"] = str(code)
        zipcode["street"] = street
        zipcode["lat"] = lat
        zipcode["long"] = long
        tx.create(zipcode)
        tx.commit()

    def Address(self, id, number):
        tx = self.__graph.begin()

        address = Node("Address")
        address["id"] = str(id)
        address["number"] = str(number)
        tx.create(address)
        tx.commit()

    def Patient(self, cpf, name, birthdate, gender, bloodtype, email):
        tx = self.__graph.begin()

        patient = Node("Patient")
        patient["cpf"] = str(cpf)
        patient["name"] = name
        patient["birthdate"] = birthdate
        patient["gender"] = gender
        patient["bloodtype"] = bloodtype
        patient["email"] = email
        tx.create(patient)
        tx.commit()

    def HasArea(self, zipcode, area):
        tx = self.__graph.begin()
        tx.run("MATCH (z:Zipcode), (a:Area) WHERE z.code = '" + str(zipcode) + "' and a.name = '" + str(area) + "' CREATE (z)-[h:HAS_AREA]->(a)")
        tx.commit()

    def HasZipcode(self, address, zipcode):
        tx = self.__graph.begin()
        tx.run("MATCH (a:Address), (z:Zipcode) WHERE a.id = '" + str(address) + "' and z.code = '" + str(zipcode) + "'  CREATE (a)-[h:HAS_ZIPCODE]->(z)")
        tx.commit()

    def HasAddress(self, patient, address):
        tx = self.__graph.begin()
        tx.run("MATCH (p:Patient), (a:Address) WHERE p.cpf = '" + str(patient) + "' and a.id = '" + str(address) + "' CREATE (p)-[h:HAS_ADDRESS]->(a)")
        tx.commit()

if __name__ == "__main__":
    conectar = ConectarNeo4J()

    conectar.limpar()

    workbook = openpyxl.load_workbook("graphData.xlsx", True)

    sheet = workbook['area']
    for i in range(2, sheet.max_row + 1):
        name = sheet["A" + str(i)].value
        city = sheet["B" + str(i)].value
        state = sheet["C" + str(i)].value
        conectar.Area(name, city, state)

    sheet = workbook['zipcode']
    for i in range(2, sheet.max_row + 1):
        code = sheet["A" + str(i)].value
        street = sheet["B" + str(i)].value
        area = sheet["C" + str(i)].value
        lat = sheet["D" + str(i)].value
        long = sheet["E" + str(i)].value
        conectar.Zipcode(code, street, lat, long)
        conectar.HasArea(code, area)

    sheet = workbook['address']
    for i in range(2, sheet.max_row + 1):
        id = sheet["A"+str(i)].value
        number = sheet["B"+str(i)].value
        zipcode = sheet["C"+str(i)].value
        conectar.Address(id, number)
        conectar.HasZipcode(id, zipcode)

    sheet = workbook['patient']
    for i in range(2, sheet.max_row + 1):
        cpf = sheet["A"+str(i)].value
        name = sheet["B"+str(i)].value
        birthdate = sheet["C"+str(i)].value
        gender = sheet["D"+str(i)].value
        bloodtype = sheet["E"+str(i)].value
        email = sheet["F"+str(i)].value
        address = sheet["G"+str(i)].value
        conectar.Patient(cpf, name, birthdate, gender, bloodtype, email)
        conectar.HasAddress(cpf, address)
